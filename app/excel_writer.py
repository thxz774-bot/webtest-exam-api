from datetime import datetime
from pathlib import Path
import json

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from template_layout import LAYOUT


def _extrair_coluna_linha(celula):
    """
    Extrai coluna e linha de uma célula (ex: "A5" -> ("A", 5)).
    
    Args:
        celula: String da célula
        
    Returns:
        Tupla (coluna, linha)
    """
    i = 0
    while i < len(celula) and celula[i].isalpha():
        i += 1
    
    coluna = celula[:i]
    linha = int(celula[i:])
    return coluna, linha


def carregar_mapeamento_gabarito(caminho):
    """Carrega o mapeamento de gabarito do JSON."""
    with open(caminho, 'r', encoding='utf-8') as f:
        return json.load(f)


def obter_carteiras_alunos_por_serie(numero_sala, mapeamento_gabarito):
    """
    Extrai apenas as carteiras (células com alunos) do mapeamento,
    organizadas por série.
    
    Args:
        numero_sala: Número da sala (13, 14, 16, 17, 18, 19, 20)
        mapeamento_gabarito: Dados do JSON de mapeamento
        
    Returns:
        Dict com carteiras por série: {"1º ano": [...], "2º ano": [...], "3º ano": [...]}
    """
    for sala_data in mapeamento_gabarito:
        nome_sala = sala_data['sala']
        if f"Sala {numero_sala}" == nome_sala:
            carteiras_por_serie = {}
            
            for posicao in sala_data['posicoes_mapeadas']:
                if posicao['tipo'] == 'CARTEIRA' and posicao['serie_obrigatoria']:
                    serie = posicao['serie_obrigatoria']
                    if serie not in carteiras_por_serie:
                        carteiras_por_serie[serie] = []
                    
                    carteiras_por_serie[serie].append({
                        'celula': posicao['celula'],
                        'serie': serie,
                        'cor': posicao['cor_obrigatoria']
                    })
            
            return carteiras_por_serie
    
    return {}


def preencher_mapa_visual(ws, numero_sala, alunos, mapeamento_gabarito):
    """
    Preenche o mapa visual de sala (células coloridas) com nomes dos alunos.
    Respeita as cores e séries obrigatórias do gabarito.
    
    Args:
        ws: Worksheet do Excel
        numero_sala: Número da sala
        alunos: Dicionário com alunos por série {"1": [...], "2": [...], "3": [...]}
        mapeamento_gabarito: Dados do JSON de mapeamento
    """
    carteiras_por_serie = obter_carteiras_alunos_por_serie(numero_sala, mapeamento_gabarito)
    
    if not carteiras_por_serie:
        return

    # Preencher cada série
    for serie_nome in ["1", "2", "3"]:
        lista_alunos = alunos.get(serie_nome, [])
        serie_completa = f"{serie_nome}º ano"
        
        if serie_completa not in carteiras_por_serie:
            continue
        
        carteiras = carteiras_por_serie[serie_completa]
        
        # Preencher cada aluno em sua carteira
        for indice, aluno in enumerate(lista_alunos):
            if indice < len(carteiras):
                celula_numero = carteiras[indice]['celula']
                
                # Extrair coluna e linha
                coluna, linha = _extrair_coluna_linha(celula_numero)
                
                # Escrever nome do aluno na célula
                _safe_set(ws, coluna, linha, aluno.nome)


def preencher_mapa(ws, sala_cfg, alunos):
    """Função original - preenche lista lateral de alunos"""
    inicio = sala_cfg["inicio"]
    fim = sala_cfg["fim"]

    todos = (
        alunos["1"] +
        alunos["2"] +
        alunos["3"]
    )

    capacidade = fim - inicio + 1

    if len(todos) > capacidade:
        raise Exception(
            f"Sala suporta {capacidade} "
            f"alunos e recebeu "
            f"{len(todos)}"
        )

    for indice, aluno in enumerate(todos):
        linha = inicio + indice
        _safe_set(ws, "AA", linha, aluno.nome)


def preencher_lista_lateral(ws, sala_cfg, alunos):
    """Função original - preenche lista lateral por série"""
    listas = sala_cfg["lista"]

    for ano in ["1", "2", "3"]:
        linha_inicio = listas[ano]["inicio"]

        for indice, aluno in enumerate(alunos[ano]):
            linha = linha_inicio + indice
            _safe_set(ws, "B", linha, aluno.nome)


def _safe_set(ws, col_letter, row, value):
    """Write `value` into the top-left cell of any merged range that contains
    the target cell (col_letter,row). If it's not merged, write directly.
    """
    try:
        col_idx = column_index_from_string(col_letter)
    except ValueError:
        print(f"Aviso: Coluna inválida '{col_letter}'")
        return

    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col_idx <= merged.max_col:
            target = ws.cell(row=merged.min_row, column=merged.min_col)
            target.value = value
            return

    ws[f"{col_letter}{row}"] = value


def gerar_excel(
    template,
    output,
    distribuicao,
    caminho_mapeamento_gabarito=None
):
    """
    Gera arquivo Excel com alunos distribuídos nas salas.
    
    Args:
        template: Caminho do arquivo template
        output: Caminho do arquivo de saída
        distribuicao: Distribuição de alunos por sala
        caminho_mapeamento_gabarito: Caminho para arquivo JSON de mapeamento de gabarito
    """
    wb = load_workbook(template)
    ws = wb["mod 1"]

    # Carregar mapeamento de gabarito se fornecido
    mapeamento_gabarito = None
    if caminho_mapeamento_gabarito and Path(caminho_mapeamento_gabarito).exists():
        try:
            mapeamento_gabarito = carregar_mapeamento_gabarito(caminho_mapeamento_gabarito)
        except Exception as e:
            print(f"Aviso: Não foi possível carregar mapeamento de gabarito: {e}")

    for sala, alunos in distribuicao.items():
        cfg = LAYOUT[str(sala)]

        # Preencher mapa visual se mapeamento de gabarito disponível
        if mapeamento_gabarito:
            preencher_mapa_visual(ws, sala, alunos, mapeamento_gabarito)

        # SEMPRE preencher as funções originais
        preencher_mapa(ws, cfg, alunos)
        preencher_lista_lateral(ws, cfg, alunos)

    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb.save(output_path)
        return output_path
    except PermissionError as exc:
        fallback_path = output_path.with_name(
            f"{output_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{output_path.suffix}"
        )
        try:
            wb.save(fallback_path)
            return fallback_path
        except PermissionError:
            raise PermissionError(
                f"Não foi possível salvar em '{output_path}' porque o arquivo está em uso ou protegido. "
                f"Feche o arquivo e execute novamente."
            ) from exc
